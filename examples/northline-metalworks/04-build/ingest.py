"""
Ingest Northline's six Excel workbooks into the relational schema.

This is the step that turns the messy "before" spreadsheets into joinable data. It handles the
real-world mess on the way in: inconsistent units (foot -> ft, each -> ea), a blank-cost catch-all,
stray blank rows, and quote line items spread across per-quote detail tabs.
"""
import pathlib

from openpyxl import load_workbook

HERE = pathlib.Path(__file__).parent
DATA = HERE.parent / "02-source-data"

UNIT_FIX = {"foot": "ft", "feet": "ft", "each": "ea"}


def _rows(ws, start=3):
    """Data rows (skip the merged title row and the header row)."""
    for r in range(start, ws.max_row + 1):
        yield [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]


def norm_unit(u):
    if u is None:
        return None
    return UNIT_FIX.get(str(u).strip().lower(), str(u).strip())


def ingest(conn):
    counts = {}

    # Materials
    ws = load_workbook(DATA / "materials-price-list.xlsx", data_only=True).active
    n = 0
    for row in _rows(ws):
        code = row[0]
        if not code:  # the stray blank row before MISC
            continue
        conn.execute(
            "INSERT INTO materials (code, description, category, unit, cost_per_unit, vendor, lead_days) VALUES (?,?,?,?,?,?,?)",
            (code, row[1], row[2], norm_unit(row[4]), row[5], row[6], row[7]),
        )
        n += 1
    counts["materials"] = n

    # Labor stations
    ws = load_workbook(DATA / "labor-rates.xlsx", data_only=True).active
    n = 0
    for row in _rows(ws):
        station = row[0]
        if not station:  # trailing note row
            continue
        base, fixed_oh, var_oh = row[1], row[2], row[3]
        conn.execute(
            "INSERT INTO labor_ops (station, base_rate, oh_rate) VALUES (?,?,?)",
            (station, base, (fixed_oh or 0) + (var_oh or 0)),
        )
        n += 1
    counts["labor_ops"] = n

    # Clients
    ws = load_workbook(DATA / "customers.xlsx", data_only=True).active
    n = 0
    for row in _rows(ws):
        if not row[0]:
            continue
        conn.execute("INSERT INTO clients (name, contact, terms) VALUES (?,?,?)", (row[0], row[1], row[3]))
        n += 1
    counts["clients"] = n

    conn.commit()

    # Quote headers (Quote Log)
    wb = load_workbook(DATA / "job-quotes-2025.xlsx", data_only=True)
    log = wb["Quote Log"]
    n = 0
    for row in _rows(log):
        number = row[0]
        if not number:
            continue
        client = conn.execute("SELECT id FROM clients WHERE name = ?", (row[2],)).fetchone()
        markup = float(str(row[7]).replace("%", "")) / 100 if row[7] is not None else 0.0
        conn.execute(
            "INSERT INTO quotes (number, client_id, description, markup, recorded_sell, status) VALUES (?,?,?,?,?,?)",
            (number, client["id"] if client else None, row[3], markup, row[8], row[9]),
        )
        n += 1
    counts["quotes"] = n
    conn.commit()

    # Quote line items (per-quote detail tabs), matched by known codes / stations, so the
    # totals block and section headers are ignored automatically.
    codes = {r["code"] for r in conn.execute("SELECT code FROM materials")}
    stations = {r["station"] for r in conn.execute("SELECT station FROM labor_ops")}
    mat_ids = {r["code"]: r["id"] for r in conn.execute("SELECT code, id FROM materials")}
    mat_lines = lab_lines = 0
    for name in wb.sheetnames:
        if name == "Quote Log":
            continue
        q = conn.execute("SELECT id FROM quotes WHERE number = ?", (name,)).fetchone()
        if not q:
            continue
        ws = wb[name]
        for r in range(1, ws.max_row + 1):
            c1 = ws.cell(row=r, column=1).value
            if c1 in codes:
                qty = ws.cell(row=r, column=3).value
                conn.execute("INSERT INTO quote_materials (quote_id, material_id, qty) VALUES (?,?,?)", (q["id"], mat_ids[c1], qty))
                mat_lines += 1
            elif c1 in stations:
                hours = ws.cell(row=r, column=2).value
                conn.execute("INSERT INTO quote_labor (quote_id, station, hours) VALUES (?,?,?)", (q["id"], c1, hours))
                lab_lines += 1
    counts["quote_materials"] = mat_lines
    counts["quote_labor"] = lab_lines

    # Production actuals
    ws = load_workbook(DATA / "production-log.xlsx", data_only=True).active
    n = 0
    for row in _rows(ws):
        job, station = row[0], row[1]
        if not job or not station:
            continue
        if station == "MATERIAL $":
            conn.execute(
                "INSERT INTO production_actuals (quote_number, station, est_hours, actual_hours, material_actual) VALUES (?,?,?,?,?)",
                (job, station, None, None, row[3]),
            )
        else:
            conn.execute(
                "INSERT INTO production_actuals (quote_number, station, est_hours, actual_hours, material_actual) VALUES (?,?,?,?,?)",
                (job, station, row[2], row[3], None),
            )
        n += 1
    counts["production_actuals"] = n

    conn.commit()
    return counts
