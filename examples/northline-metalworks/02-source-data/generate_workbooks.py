#!/usr/bin/env python3
"""
Generate the synthetic "before" spreadsheets for Northline Metalworks (sample shop #1).

Everything here is invented for demonstration. No real company, vendor, customer, or price.
The data is deliberately a little messy (inconsistent units, a blank cost, a hand-overridden
sell price) so the "after" story is honest. Values are baked (not live formulas) so the built
system can ingest them cleanly and the output is reproducible.

Run:  python3 generate_workbooks.py
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT = Path(__file__).parent

TITLE_FILL = PatternFill("solid", fgColor="1F3B4D")
HEADER_FILL = PatternFill("solid", fgColor="DDE3E8")
TITLE_FONT = Font(bold=True, color="FFFFFF", size=12)
HEADER_FONT = Font(bold=True)


def write_sheet(ws, title, headers, rows, widths=None):
    """Merged title row, a bold header row, then data. Mirrors how a shop's tab actually looks."""
    ncol = len(headers)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    tcell = ws.cell(row=1, column=1, value=title)
    tcell.fill = TITLE_FILL
    tcell.font = TITLE_FONT
    tcell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for r, row in enumerate(rows, start=3):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)

    for c in range(1, ncol + 1):
        w = (widths[c - 1] if widths and c - 1 < len(widths) else 16)
        ws.column_dimensions[get_column_letter(c)].width = w


# ── Reference data (the shop's real-world facts, invented) ──────────────────────────────

# code: [description, category, size/spec, unit, cost_per_unit, vendor, lead_days, last_updated, notes]
MATERIALS = {
    "ST-TUBE-2020-11": ["Steel tube 2x2 11ga", "Tube", '2" x 2" 11ga', "ft", 4.20, "Midwest Steel Supply", 5, "2025-09-14", ""],
    "ST-TUBE-1515-14": ["Steel tube 1.5x1.5 14ga", "Tube", '1.5" x 1.5" 14ga', "ft", 2.65, "Midwest Steel Supply", 5, "2025-09-14", ""],
    "ST-ANGLE-2020-025": ["Steel angle 2x2 1/4", "Angle", '2" x 2" x 1/4"', "foot", 3.10, "Midwest Steel Supply", 5, "2025-08-02", "unit says foot, everywhere else ft"],
    "ST-SHEET-14": ["Steel sheet 14ga 4x8", "Sheet", "4' x 8' 14ga", "sheet", 92.00, "Midwest Steel Supply", 7, "2025-09-14", ""],
    "ST-PLATE-025": ["Steel plate 1/4 4x8", "Plate", '4\' x 8\' 1/4"', "sheet", 214.00, "Midwest Steel Supply", 10, "2025-07-19", ""],
    "HW-BOLT-38": ["Bolt 3/8-16 x 1 grade 5", "Fastener", '3/8-16 x 1"', "ea", 0.34, "Fastenal", 3, "2025-06-30", ""],
    "HW-NUT-38": ["Nut 3/8-16", "Fastener", "3/8-16", "each", 0.11, "Fastenal", 3, "2025-06-30", "unit says each"],
    "HW-FOOT-LEVEL": ["Leveling foot 3/8 stem", "Fastener", "3/8 stem", "ea", 1.05, "Fastenal", 4, "", "no last-updated date"],
    "WELD-WIRE-035": ["MIG wire .035", "Consumable", "33 lb spool", "lb", 2.80, "Midwest Steel Supply", 5, "2025-05-11", ""],
    "PC-BLACK": ["Powder coat black", "Finish", "per sq ft", "sqft", 0.48, "(in-house)", 0, "2025-01-20", ""],
    "MISC": ["Misc / shop supplies", "Misc", "-", "lot", None, "", None, "", "catch-all, never itemized, no cost entered"],
}

# station: [base_rate, fixed_oh, var_oh]
LABOR = {
    "Saw / Cut": [22.0, 16.0, 8.0],
    "Brake / Form": [24.0, 16.0, 8.0],
    "Weld": [28.0, 18.0, 10.0],
    "Grind / Finish": [22.0, 15.0, 7.0],
    "Powder Coat": [20.0, 14.0, 6.0],
    "Assembly": [20.0, 14.0, 6.0],
    "Delivery": [22.0, 12.0, 6.0],
}

# code: [name, contact, phone, terms, notes]
CUSTOMERS = [
    ["Cedar Ridge Contractors", "M. Alvarez", "555-0142", "Net 30", "repeat shelving buyer"],
    ["Fulton Warehouse Co", "D. Okafor", "555-0187", "Net 30", "pallet racking, quarterly"],
    ["Bright Line Interiors", "S. Kim", "555-0119", "Due on receipt", "custom, design-led"],
    ["Halstead Grocery Group", "J. Patel", "555-0164", "Net 45", "slow pay, big orders"],
    ["Prairie Steel Racking", "R. Boone", "555-0173", "Net 30", "reseller, price sensitive"],
    ["Delco Auto Parts", "T. Nguyen", "555-0128", "Net 30", "wall shelving"],
    ["Northside Storage", "L. Freeman", "555-0155", "COD", "one-offs"],
    ["Iron & Oak Furnishings", "P. Costa", "555-0190", "Net 30", "brackets, table bases"],
]

# Quotes: structured line items so totals stay internally consistent.
# materials: list of (code, qty). labor: list of (station, hours).
QUOTES = [
    {
        "num": "Q-2025-018", "date": "2025-09-22", "customer": "Cedar Ridge Contractors",
        "desc": "Boltless shelving unit 48x18x72, 5 shelf, qty 10", "qty": 10, "markup": 0.35, "status": "Won",
        "materials": [
            ("ST-TUBE-1515-14", 240), ("ST-ANGLE-2020-025", 200), ("ST-SHEET-14", 5),
            ("HW-BOLT-38", 160), ("HW-NUT-38", 160), ("HW-FOOT-LEVEL", 40), ("PC-BLACK", 300),
        ],
        "labor": [("Saw / Cut", 4.0), ("Brake / Form", 2.0), ("Weld", 6.0), ("Grind / Finish", 3.0), ("Powder Coat", 3.0), ("Assembly", 5.0)],
        "override_sell": None,
    },
    {
        "num": "Q-2025-023", "date": "2025-10-06", "customer": "Fulton Warehouse Co",
        "desc": "Welded pallet rack upright 8ft, qty 6", "qty": 6, "markup": 0.35, "status": "Won",
        "materials": [
            ("ST-TUBE-2020-11", 96), ("ST-ANGLE-2020-025", 144), ("ST-PLATE-025", 1.5),
            ("HW-BOLT-38", 48), ("WELD-WIRE-035", 12), ("PC-BLACK", 240),
        ],
        "labor": [("Saw / Cut", 3.0), ("Weld", 8.4), ("Grind / Finish", 2.4), ("Powder Coat", 2.4), ("Assembly", 1.2)],
        "override_sell": None,
    },
    {
        "num": "Q-2025-031", "date": "2025-10-20", "customer": "Bright Line Interiors",
        "desc": "Custom steel handrail, lobby (one-off)", "qty": 1, "markup": 0.35, "status": "Won",
        "materials": [
            ("ST-TUBE-1515-14", 40), ("ST-PLATE-025", 0.25), ("HW-BOLT-38", 12), ("WELD-WIRE-035", 2), ("PC-BLACK", 60),
        ],
        "labor": [("Saw / Cut", 1.0), ("Brake / Form", 0.5), ("Weld", 3.0), ("Grind / Finish", 1.5), ("Powder Coat", 0.5), ("Assembly", 0.5), ("Delivery", 1.0)],
        "override_sell": 750.00,  # owner typed a round number over the estimate; see notes
    },
]

# Cut lists (part, material code, length/size, qty) per job.
CUT_LISTS = {
    "Q-2025-018": [
        ["Corner post", "ST-TUBE-1515-14", "72 in", 40],
        ["Shelf frame long", "ST-ANGLE-2020-025", "46 in", 100],
        ["Shelf frame short", "ST-ANGLE-2020-025", "16 in", 100],
        ["Shelf panel", "ST-SHEET-14", "48 x 18 in", 50],
    ],
    "Q-2025-023": [
        ["Upright post", "ST-TUBE-2020-11", "96 in", 12],
        ["Diagonal brace", "ST-ANGLE-2020-025", "24 in", 72],
        ["Base plate", "ST-PLATE-025", "6 x 6 in", 12],
    ],
}

# Production actuals (hand-logged) for the jobs that were actually built.
# station: [est_hours, actual_hours]
ACTUALS = {
    "Q-2025-018": {
        "material_est": None, "material_actual": 2075.00,  # est filled from computed below
        "labor": {"Saw / Cut": [4.0, 4.2], "Brake / Form": [2.0, 2.0], "Weld": [6.0, 8.0], "Grind / Finish": [3.0, 3.5], "Powder Coat": [3.0, 3.0], "Assembly": [5.0, 5.2]},
        "note": "welding ran well over; job felt tight and nobody knew why",
    },
    "Q-2025-023": {
        "material_est": None, "material_actual": 1360.00,
        "labor": {"Saw / Cut": [3.0, 3.0], "Weld": [8.4, 8.6], "Grind / Finish": [2.4, 2.4], "Powder Coat": [2.4, 2.5], "Assembly": [1.2, 1.2]},
        "note": "close to estimate",
    },
}


# ── Compute quote totals from the line items (keeps everything consistent) ──────────────

def quote_totals(q):
    material = 0.0
    for code, qty in q["materials"]:
        cost = MATERIALS[code][4] or 0.0
        material += cost * qty
    labor = 0.0
    overhead = 0.0
    for station, hours in q["labor"]:
        base, fixed_oh, var_oh = LABOR[station]
        labor += hours * base
        overhead += hours * (fixed_oh + var_oh)
    subtotal = material + labor + overhead
    computed_sell = round(subtotal * (1 + q["markup"]), 2)
    sell = q["override_sell"] if q["override_sell"] is not None else computed_sell
    return {
        "material": round(material, 2), "labor": round(labor, 2), "overhead": round(overhead, 2),
        "computed_sell": computed_sell, "sell": round(sell, 2),
    }


# ── Workbook writers ────────────────────────────────────────────────────────────────────

def build_materials():
    wb = Workbook()
    ws = wb.active
    ws.title = "Price List"
    headers = ["Item Code", "Description", "Category", "Size / Spec", "Unit", "Cost/Unit", "Vendor", "Lead (days)", "Last Updated", "Notes"]
    rows = []
    codes = list(MATERIALS.keys())
    for code in codes:
        if code == "MISC":
            rows.append([None] * len(headers))  # stray blank row before the catch-all, as in the real file
        d = MATERIALS[code]
        rows.append([code, d[0], d[1], d[2], d[3], d[4], d[5], d[6], d[7], d[8]])
    write_sheet(ws, "NORTHLINE METALWORKS  ·  MATERIALS PRICE LIST (updated as invoices come in)", headers, rows,
                widths=[18, 26, 12, 18, 8, 10, 22, 11, 14, 34])
    wb.save(OUT / "materials-price-list.xlsx")


def build_labor():
    wb = Workbook()
    ws = wb.active
    ws.title = "Shop Rates"
    headers = ["Station", "Base Labor $/hr", "Fixed OH $/hr", "Variable OH $/hr", "Shop Rate $/hr", "Notes"]
    rows = []
    for station, (base, fixed_oh, var_oh) in LABOR.items():
        rows.append([station, base, fixed_oh, var_oh, round(base + fixed_oh + var_oh, 2), ""])
    rows.append(["", "", "", "", "", "set sometime last year; not revisited since"])
    write_sheet(ws, "NORTHLINE METALWORKS  ·  LABOR & OVERHEAD RATES", headers, rows,
                widths=[16, 16, 15, 17, 15, 40])
    wb.save(OUT / "labor-rates.xlsx")


def build_quotes():
    wb = Workbook()
    log = wb.active
    log.title = "Quote Log"
    headers = ["Quote #", "Date", "Customer", "Description", "Material $", "Labor $", "Overhead $", "Markup %", "Sell $", "Status", "Notes"]
    rows = []
    for q in QUOTES:
        t = quote_totals(q)
        note = ""
        if q["override_sell"] is not None:
            note = f"sell set by hand to {q['override_sell']:.0f}; formula gives {t['computed_sell']:.2f}"
        rows.append([q["num"], q["date"], q["customer"], q["desc"], t["material"], t["labor"], t["overhead"],
                     f"{int(q['markup']*100)}%", t["sell"], q["status"], note])
    write_sheet(log, "NORTHLINE METALWORKS  ·  2025 QUOTE LOG", headers, rows,
                widths=[12, 11, 24, 40, 11, 10, 11, 9, 10, 8, 40])

    # One detail tab per quote (materials block, then labor block).
    for q in QUOTES:
        ws = wb.create_sheet(q["num"])
        ws.merge_cells("A1:F1")
        c = ws.cell(row=1, column=1, value=f"{q['num']}  ·  {q['customer']}  ·  {q['desc']}")
        c.fill = TITLE_FILL
        c.font = TITLE_FONT
        r = 3
        ws.cell(row=r, column=1, value="MATERIALS").font = HEADER_FONT
        r += 1
        for h in ["Item Code", "Description", "Qty", "Unit", "Unit Cost", "Amount"]:
            ws.cell(row=r, column=["Item Code", "Description", "Qty", "Unit", "Unit Cost", "Amount"].index(h) + 1, value=h).font = HEADER_FONT
        r += 1
        for code, qty in q["materials"]:
            d = MATERIALS[code]
            cost = d[4] or 0.0
            ws.cell(row=r, column=1, value=code)
            ws.cell(row=r, column=2, value=d[0])
            ws.cell(row=r, column=3, value=qty)
            ws.cell(row=r, column=4, value=d[3])
            ws.cell(row=r, column=5, value=cost)
            ws.cell(row=r, column=6, value=round(cost * qty, 2))
            r += 1
        r += 1
        ws.cell(row=r, column=1, value="LABOR").font = HEADER_FONT
        r += 1
        for h in ["Station", "Hours", "Base $/hr", "OH $/hr", "Labor $", "Overhead $"]:
            ws.cell(row=r, column=["Station", "Hours", "Base $/hr", "OH $/hr", "Labor $", "Overhead $"].index(h) + 1, value=h).font = HEADER_FONT
        r += 1
        for station, hours in q["labor"]:
            base, fixed_oh, var_oh = LABOR[station]
            ws.cell(row=r, column=1, value=station)
            ws.cell(row=r, column=2, value=hours)
            ws.cell(row=r, column=3, value=base)
            ws.cell(row=r, column=4, value=round(fixed_oh + var_oh, 2))
            ws.cell(row=r, column=5, value=round(hours * base, 2))
            ws.cell(row=r, column=6, value=round(hours * (fixed_oh + var_oh), 2))
            r += 1
        t = quote_totals(q)
        r += 1
        for label, val in [("Material $", t["material"]), ("Labor $", t["labor"]), ("Overhead $", t["overhead"]),
                           (f"Markup {int(q['markup']*100)}%", ""), ("SELL $", t["sell"])]:
            ws.cell(row=r, column=5, value=label).font = HEADER_FONT
            ws.cell(row=r, column=6, value=val)
            r += 1
        if q["override_sell"] is not None:
            ws.cell(row=r + 1, column=1, value=f"(sell hand-set to {q['override_sell']:.0f}; formula gives {t['computed_sell']:.2f})")
        for col, w in zip("ABCDEF", [18, 26, 10, 10, 12, 12]):
            ws.column_dimensions[col].width = w

    wb.save(OUT / "job-quotes-2025.xlsx")


def build_cut_lists():
    wb = Workbook()
    first = True
    for job, parts in CUT_LISTS.items():
        ws = wb.active if first else wb.create_sheet()
        ws.title = job
        first = False
        headers = ["Part", "Material Code", "Length / Size", "Qty", "Notes"]
        write_sheet(ws, f"CUT LIST  ·  {job}", headers, [[p[0], p[1], p[2], p[3], ""] for p in parts],
                    widths=[22, 20, 16, 8, 30])
    wb.save(OUT / "cut-lists.xlsx")


def build_customers():
    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"
    headers = ["Customer", "Contact", "Phone", "Terms", "Notes"]
    write_sheet(ws, "NORTHLINE METALWORKS  ·  CUSTOMERS", headers, CUSTOMERS, widths=[26, 14, 12, 16, 32])
    wb.save(OUT / "customers.xlsx")


def build_production_log():
    wb = Workbook()
    ws = wb.active
    ws.title = "Job Log"
    headers = ["Job #", "Station", "Est Hours", "Actual Hours", "Hours Var", "Note"]
    rows = []
    for job, data in ACTUALS.items():
        for station, (est, act) in data["labor"].items():
            rows.append([job, station, est, act, round(act - est, 2), ""])
        rows.append([job, "MATERIAL $", round(quote_totals(next(q for q in QUOTES if q["num"] == job))["material"], 2),
                     data["material_actual"], "", data["note"]])
        rows.append([None] * len(headers))
    write_sheet(ws, "NORTHLINE METALWORKS  ·  PRODUCTION LOG (hand-entered, rarely reviewed)", headers, rows,
                widths=[12, 16, 11, 13, 11, 40])
    wb.save(OUT / "production-log.xlsx")


def main():
    build_materials()
    build_labor()
    build_quotes()
    build_cut_lists()
    build_customers()
    build_production_log()
    print("Wrote workbooks to", OUT)
    for q in QUOTES:
        t = quote_totals(q)
        print(f"  {q['num']}: material {t['material']}, labor {t['labor']}, overhead {t['overhead']}, sell {t['sell']} (formula {t['computed_sell']})")


if __name__ == "__main__":
    main()
